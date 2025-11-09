#!/usr/bin/env python3
import openpyxl
import sys

# Load with formulas
wb_formulas = openpyxl.load_workbook('/Users/ryder/Downloads/Relaxed-Moon-Avoidance.xlsx', data_only=False)
ws_formulas = wb_formulas.active

# Load with calculated values
wb_values = openpyxl.load_workbook('/Users/ryder/Downloads/Relaxed-Moon-Avoidance.xlsx', data_only=True)
ws_values = wb_values.active

print("=== PARAMETERS ===")
print(f"Separation (Classic): {ws_values.cell(3, 2).value}")
print(f"Width (Classic): {ws_values.cell(4, 2).value}")
print(f"Separation (Relaxed): {ws_values.cell(3, 5).value}")
print(f"Width (Relaxed): {ws_values.cell(4, 5).value}")
print(f"Altitude: {ws_values.cell(5, 5).value}")
print(f"Relax Scale: {ws_values.cell(6, 5).value}")
print(f"Min Alt: {ws_values.cell(7, 5).value}")
print(f"Max Alt: {ws_values.cell(8, 5).value}")
print(f"Lunar Cycle: {ws_values.cell(3, 7).value}")

print("\n=== FORMULAS ===")
print(f"Classic formula (C10): {ws_formulas.cell(10, 3).value}")
print(f"Relaxed formula (D10): {ws_formulas.cell(10, 4).value}")

print("\n=== SAMPLE VALUES ===")
for row in [10, 15, 20, 25, 30]:
    age = ws_values.cell(row, 2).value
    classic = ws_values.cell(row, 3).value
    relaxed = ws_values.cell(row, 4).value
    if isinstance(classic, (int, float)) and isinstance(relaxed, (int, float)):
        print(f"Age {age}: Classic={classic:.2f}, Relaxed={relaxed:.2f}")
    else:
        print(f"Age {age}: Classic={classic}, Relaxed={relaxed}")

print("\n=== FULL MOON CHECK (Age 15) ===")
age_15 = ws_values.cell(15, 2).value
classic_15 = ws_values.cell(15, 3).value
relaxed_15 = ws_values.cell(15, 4).value
if isinstance(classic_15, (int, float)) and isinstance(relaxed_15, (int, float)):
    print(f"Age {age_15}: Classic={classic_15:.2f}, Relaxed={relaxed_15:.2f}")
else:
    print(f"Age {age_15}: Classic={classic_15}, Relaxed={relaxed_15}")

print("\n=== KEY POINTS ===")
print("Age 0 (new moon):")
age_0 = ws_values.cell(10, 2).value
classic_0 = ws_values.cell(10, 3).value
relaxed_0 = ws_values.cell(10, 4).value
if isinstance(classic_0, (int, float)) and isinstance(relaxed_0, (int, float)):
    print(f"  Classic={classic_0:.2f}, Relaxed={relaxed_0:.2f}")

print("Age 15 (full moon):")
if isinstance(classic_15, (int, float)) and isinstance(relaxed_15, (int, float)):
    print(f"  Classic={classic_15:.2f}, Relaxed={relaxed_15:.2f}")

age_29 = ws_values.cell(39, 2).value
classic_29 = ws_values.cell(39, 3).value
relaxed_29 = ws_values.cell(39, 4).value
print("Age 29 (next new moon):")
if isinstance(classic_29, (int, float)) and isinstance(relaxed_29, (int, float)):
    print(f"  Classic={classic_29:.2f}, Relaxed={relaxed_29:.2f}")

